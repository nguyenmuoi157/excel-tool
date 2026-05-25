import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
from db_manager import DBManager
from exporter import Exporter

class DataProcessorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Project X: Large Data Filter & Export (Tkinter Version)")
        self.geometry("1400x900")

        self.db = DBManager()
        self.exporter = Exporter()

        self.current_page = 1
        self.page_size = 50
        self.total_records = 0
        self.filter_rows = []
        self.current_columns = None

        self._detect_dark_mode()
        self._init_ui()
        self._check_existing_data()

    def _detect_dark_mode(self):
        """Detect macOS dark mode for theme-aware colors."""
        import subprocess, platform
        self._is_dark = False
        if platform.system() == 'Darwin':
            try:
                out = subprocess.check_output(
                    'defaults read -g AppleInterfaceStyle 2>/dev/null',
                    shell=True, text=True
                ).strip()
                self._is_dark = out.lower() == 'dark'
            except Exception:
                pass
        # Status bar colors
        self._status_bg = '#1e1e1e' if self._is_dark else '#f0f0f0'
        self._status_fg = '#ffffff' if self._is_dark else '#000000'
        self._flash_bg = '#1b5e20' if self._is_dark else '#e8f5e9'
        self._flash_fg = '#ffffff' if self._is_dark else '#2e7d32'

    def _init_ui(self):
        # Header
        header_frame = ttk.Frame(self, padding=10)
        header_frame.pack(fill=tk.X)

        ttk.Label(header_frame, text="Large Data Processor", font=("Helvetica", 16, "bold")).pack(side=tk.LEFT)

        btn_frame = ttk.Frame(header_frame)
        btn_frame.pack(side=tk.RIGHT)

        self.btn_pick = ttk.Button(btn_frame, text="Chọn file Excel/CSV", command=self.pick_file)
        self.btn_pick.pack(side=tk.LEFT, padx=5)
        self.btn_export = ttk.Button(btn_frame, text="Xuất Excel", command=self.export_file)
        self.btn_export.pack(side=tk.LEFT, padx=5)

        # Status & Progress
        self.status_var = tk.StringVar(value="Sẵn sàng")
        self.progress_var = tk.DoubleVar(value=0)

        status_frame = tk.Frame(self, background=self._status_bg, padx=10, pady=2)
        status_frame.pack(fill=tk.X)
        self.status_bar = status_frame
        tk.Label(status_frame, textvariable=self.status_var, background=self._status_bg,
                 foreground=self._status_fg, anchor='w').pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.progress_bar = ttk.Progressbar(status_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=10)

        # Filters Area (Dynamic)
        self.filter_frame_container = ttk.LabelFrame(self, text="Bộ lọc nâng cao", padding=10)
        self.filter_frame_container.pack(fill=tk.X, padx=10, pady=5)

        self.filter_list_frame = ttk.Frame(self.filter_frame_container)
        self.filter_list_frame.pack(fill=tk.X)

        filter_actions = ttk.Frame(self.filter_frame_container)
        filter_actions.pack(fill=tk.X, pady=5)
        self.btn_filter_add = ttk.Button(filter_actions, text="+ Thêm điều kiện", command=self._add_filter_row)
        self.btn_filter_add.pack(side=tk.LEFT)
        self.btn_filter_apply = ttk.Button(filter_actions, text="Áp dụng bộ lọc", command=self._apply_filters)
        self.btn_filter_apply.pack(side=tk.LEFT, padx=10)

        # Data Treeview
        tree_frame = ttk.Frame(self, padding=10)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(tree_frame, columns=("info"), show="headings")
        self.tree.heading("info", text="Thông tin")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(column=0, row=0, sticky='nsew')
        vsb.grid(column=1, row=0, sticky='ns')
        hsb.grid(column=0, row=1, sticky='ew')
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(1, weight=0)

        # Pagination
        nav_frame = ttk.Frame(self, padding=10)
        nav_frame.pack(fill=tk.X)

        center_nav = ttk.Frame(nav_frame)
        center_nav.pack(anchor=tk.CENTER)

        self.btn_prev = ttk.Button(center_nav, text="< Trước", command=self.prev_page)
        self.btn_prev.pack(side=tk.LEFT)
        self.page_label = ttk.Label(center_nav, text="Trang 1")
        self.page_label.pack(side=tk.LEFT, padx=10)
        self.btn_next = ttk.Button(center_nav, text="Sau >", command=self.next_page)
        self.btn_next.pack(side=tk.LEFT)

    def _lock_ui(self):
        """Disable all interactive widgets. Status bar + progress bar still visible."""
        for btn in [self.btn_pick, self.btn_export, self.btn_filter_add,
                    self.btn_filter_apply, self.btn_prev, self.btn_next]:
            btn.config(state="disabled")
        for col_var, op_var, val_var, frame in self.filter_rows:
            for child in frame.winfo_children():
                if isinstance(child, ttk.Combobox):
                    child.config(state="disabled")
                elif isinstance(child, ttk.Entry):
                    child.config(state="disabled")
                elif isinstance(child, ttk.Button):
                    child.config(state="disabled")

    def _flash_status(self, text, duration_ms=2000):
        """Temporarily change status bar color and text, then fade back."""
        original_bg = self.status_bar.cget("background")
        original_text = self.status_var.get()
        label = self.status_bar.winfo_children()[0]
        original_fg = label.cget("foreground")

        self.status_bar.config(background=self._flash_bg)
        label.config(background=self._flash_bg, foreground=self._flash_fg)
        self.status_var.set(text)
        self.after(duration_ms, lambda: self.status_bar.config(background=original_bg))
        self.after(duration_ms, lambda: label.config(background=original_bg, foreground=original_fg))
        self.after(duration_ms, lambda: self.status_var.set(original_text))

    def _unlock_ui(self):
        """Re-enable all interactive widgets."""
        for btn in [self.btn_pick, self.btn_export, self.btn_filter_add,
                    self.btn_filter_apply, self.btn_prev, self.btn_next]:
            btn.config(state="normal")
        for col_var, op_var, val_var, frame in self.filter_rows:
            for child in frame.winfo_children():
                if isinstance(child, ttk.Combobox):
                    child.config(state="readonly")
                elif isinstance(child, ttk.Entry):
                    child.config(state="normal")
                elif isinstance(child, ttk.Button):
                    child.config(state="normal")

    def _check_existing_data(self):
        if self.db.check_existing_db():
            self.status_var.set("Đã tải dữ liệu từ phiên trước.")
            self._apply_filters()
        else:
            self.status_var.set("Chưa có dữ liệu. Vui lòng chọn file.")

    def pick_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel/CSV Files", "*.csv *.xlsx *.xls")])
        if file_path:
            self._lock_ui()
            self.status_var.set(f"Đang chuẩn bị: {os.path.basename(file_path)}...")
            threading.Thread(target=self._process_file_thread, args=(file_path,), daemon=True).start()

    def _count_lines(self, filename):
        try:
            f = open(filename, 'rb')
            lines = 0
            buf_size = 1024 * 1024
            read_f = f.raw.read
            buf = read_f(buf_size)
            while buf:
                lines += buf.count(b'\n')
                buf = read_f(buf_size)
            f.close()
            return lines
        except:
            return 0

    def _process_file_thread(self, file_path):
        try:
            ext = os.path.splitext(file_path)[1].lower()
            total_rows = 0

            if ext == '.csv':
                self.after(0, lambda: self.status_var.set("Đang đếm số dòng..."))
                self.after(0, lambda: self.progress_bar.configure(mode='indeterminate'))
                self.after(0, self.progress_bar.start, 10)

                total_rows = self._count_lines(file_path)

                self.after(0, self.progress_bar.stop)
                self.after(0, lambda: self.progress_bar.configure(mode='determinate'))
                self.after(0, lambda: self.progress_var.set(0))

                if total_rows > 0:
                    self.after(0, lambda: self.status_var.set(f"Đang import... (Tổng: {total_rows} dòng)"))
                else:
                    self.after(0, lambda: self.status_var.set("Đang import..."))
            elif ext in ['.xlsx', '.xls']:
                import openpyxl
                self.after(0, lambda: self.status_var.set("Đang đếm số dòng..."))
                self.after(0, lambda: self.progress_bar.configure(mode='indeterminate'))
                self.after(0, self.progress_bar.start, 10)

                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                total_rows = wb.active.max_row or 0
                wb.close()
                del wb

                self.after(0, self.progress_bar.stop)
                self.after(0, lambda: self.progress_bar.configure(mode='determinate'))
                self.after(0, lambda: self.progress_var.set(0))

                if total_rows > 1:
                    self.after(0, lambda: self.status_var.set(f"Đang import... (Tổng: {total_rows} dòng)"))
                else:
                    self.after(0, lambda: self.status_var.set("Đang import file Excel..."))
            else:
                self.after(0, lambda: self.status_var.set("Đang import..."))
                self.after(0, lambda: self.progress_bar.configure(mode='indeterminate'))
                self.after(0, self.progress_bar.start, 10)

            def progress_cb(processed):
                if total_rows > 1:
                    pct = (processed / (total_rows - 1)) * 100
                    pct = min(pct, 100)
                    self.after(0, lambda p=pct: self.progress_var.set(p))
                    self.after(0, lambda p=processed: self.status_var.set(
                        f"Đã import {p:,}/{total_rows - 1:,} dòng ({int(pct)}%)"))
                else:
                    self.after(0, lambda p=processed: self.status_var.set(f"Đã import {p:,} dòng..."))

            self.db.import_file(file_path, progress_callback=progress_cb)

            self.after(0, lambda: self.status_var.set("Đang tạo index..."))
            self.db.create_indexes()
            self.after(0, self._on_import_success)

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Lỗi", str(e)))
            self.after(0, lambda: self.status_var.set("Lỗi import"))
            self.after(0, self.progress_bar.stop)
            self.after(0, self._unlock_ui)

    def _on_import_success(self):
        self.progress_bar.stop()
        self.progress_bar.configure(mode='determinate')
        self.progress_var.set(100)
        self.status_var.set("Import thành công!")

        self._clear_filters()
        self.current_page = 1
        self._unlock_ui()
        self._update_table()

    def _clear_filters(self):
        for *_, frame in self.filter_rows:
            frame.destroy()
        self.filter_rows.clear()
        self._add_filter_row()

    def _add_filter_row(self):
        cols = self.db.get_columns()
        if not cols: return

        frame = ttk.Frame(self.filter_list_frame)
        frame.pack(fill=tk.X, pady=2)

        col_var = tk.StringVar()
        col_cb = ttk.Combobox(frame, textvariable=col_var, values=cols, width=15, state="readonly")
        col_cb.set(cols[0])
        col_cb.pack(side=tk.LEFT, padx=2)

        op_var = tk.StringVar(value="contains")
        ops = ["=", ">", "<", ">=", "<=", "contains", "starts_with", "ends_with", "!="]
        op_cb = ttk.Combobox(frame, textvariable=op_var, values=ops, width=10, state="readonly")
        op_cb.pack(side=tk.LEFT, padx=2)

        val_var = tk.StringVar()
        entry = ttk.Entry(frame, textvariable=val_var, width=20)
        entry.pack(side=tk.LEFT, padx=2, fill=tk.X, expand=True)

        def delete_row(f=frame):
            f.destroy()
            self.filter_rows = [r for r in self.filter_rows if r[3] != f]

        ttk.Button(frame, text="X", width=3, command=delete_row).pack(side=tk.LEFT, padx=2)

        self.filter_rows.append((col_var, op_var, val_var, frame))

    def _apply_filters(self):
        self.current_page = 1
        filters = self._get_current_filters()

        where_clause, params = self.db._build_where_clause(filters)
        with self.db._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(f'SELECT COUNT(*) FROM "{self.db.table_name}"' + where_clause, params)
            filtered_count = cursor.fetchone()[0]

        self._update_table()

        # Flash effect on status bar to signal filter applied
        self._flash_status(f"Đã lọc: {filtered_count:,} dòng")

    def _get_current_filters(self):
        filters = []
        for col_var, op_var, val_var, _ in self.filter_rows:
            col = col_var.get()
            val = val_var.get().strip()
            if col and val:
                filters.append({
                    'col': col,
                    'op': op_var.get(),
                    'val': val
                })
        return filters

    def _update_table(self):
        cols = self.db.get_columns()

        if cols != self.current_columns:
            self.tree["columns"] = cols
            self.current_columns = cols

            for col in cols:
                self.tree.heading(col, text=col)
                header_width = len(col) * 15 + 30
                width = max(150, min(header_width, 500))
                self.tree.column(col, width=width, stretch=False, anchor='w')

        for item in self.tree.get_children():
            self.tree.delete(item)

        filters = self._get_current_filters()
        rows, total = self.db.get_data_paginated(self.current_page, self.page_size, filters)
        self.total_records = total

        for row in rows:
            self.tree.insert("", "end", values=row)

        total_pages = (total // self.page_size) + 1 if total > 0 else 1
        self.page_label.config(text=f"Trang {self.current_page} / {total_pages} (Tổng: {total})")

    def prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self._update_table()

    def next_page(self):
        if self.current_page * self.page_size < self.total_records:
            self.current_page += 1
            self._update_table()

    def export_file(self):
        output_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Lưu file Excel"
        )
        if not output_path:
            return

        self._lock_ui()
        self.status_var.set("Đang xuất file...")
        self.progress_bar['mode'] = 'determinate'
        self.progress_var.set(0)
        threading.Thread(target=self._export_thread, args=(output_path,), daemon=True).start()

    def _export_thread(self, output_path):
        try:
            filters = self._get_current_filters()
            columns = self.db.get_columns()

            # Count total rows for determinate progress
            where_clause, params = self.db._build_where_clause(filters)
            with self.db._get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(f'SELECT COUNT(*) FROM "{self.db.table_name}"' + where_clause, params)
                total_export = cursor.fetchone()[0]

            data_stream = self.db.export_filtered_data_streaming(filters)

            def progress_cb(n):
                pct = (n / total_export * 100) if total_export > 0 else 0
                pct = min(pct, 100)
                self.after(0, lambda p=pct: self.progress_var.set(p))
                pct_final = pct
                rows_final = n
                self.after(0, lambda: self.status_var.set(
                    f"Đang ghi {rows_final:,}/{total_export:,} dòng ({int(pct_final)}%)"))

            gen_total, written_total = self.exporter.export_from_generator(data_stream, columns, output_path, progress_callback=progress_cb)

            # Verify: count actual rows in exported file vs expected
            actual_data_rows = 0
            num_sheets = 0
            import openpyxl
            wb_verify = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
            for ws in wb_verify.worksheets:
                actual_data_rows += max(0, (ws.max_row or 0) - 1)
                num_sheets += 1
            wb_verify.close()

            # Debug info: show where the mismatch is
            debug_info = (f"[DEBUG] gen_total={gen_total:,}, written_total={written_total:,}, "
                          f"file_rows={actual_data_rows:,}, db_count={total_export:,}")

            if actual_data_rows != total_export or written_total != gen_total:
                self.after(0, lambda ar=actual_data_rows, te=total_export, ns=num_sheets,
                            gt=gen_total, wt=written_total:
                    messagebox.showwarning("Cảnh báo",
                        f"Số dòng không khớp!\n"
                        f"DB COUNT: {te:,} dòng\n"
                        f"Generator yielded: {gt:,} dòng\n"
                        f"Code wrote: {wt:,} dòng\n"
                        f"File thực tế: {ar:,} dòng ({ns} sheet)\n"
                        f"Thiếu: {te - ar:,} dòng"))
            else:
                self.after(0, lambda: messagebox.showinfo("Thành công",
                    f"Đã xuất file:\n{output_path}\n\n"
                    f"Tổng: {total_export:,} dòng ({num_sheets} sheet)"))

            self.after(0, self.progress_bar.stop)
            self.after(0, lambda: self.status_var.set("Xuất file hoàn tất"))
        except Exception as e:
            self.after(0, self.progress_bar.stop)
            self.after(0, lambda: messagebox.showerror("Lỗi xuất file", str(e)))
            self.after(0, lambda: self.status_var.set("Lỗi xuất file"))
        finally:
            self.after(0, self._unlock_ui)

if __name__ == "__main__":
    app = DataProcessorApp()
    app.mainloop()